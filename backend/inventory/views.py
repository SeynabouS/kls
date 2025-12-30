from __future__ import annotations

import csv
import io
import posixpath
import re
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from decimal import Decimal, InvalidOperation

from collections import defaultdict
from datetime import date
from django.core.files.base import ContentFile
from django.db import transaction as db_transaction
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rest_framework import permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from inventory.audit import log_audit_event
from inventory.models import AuditEvent, Dette, Envoi, Produit, Stock, TauxChange, Transaction
from inventory.serializers import (
    AuditEventSerializer,
    DetteSerializer,
    EnvoiSerializer,
    ProduitSerializer,
    StockSerializer,
    TauxChangeSerializer,
    TransactionSerializer,
)
from inventory.services import (
    disable_stock_recalc,
    get_current_exchange_rate,
    recalculate_stock_for_product,
)


def _csv_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return str(value.quantize(Decimal("0.01"))).replace(".", ",")
    return str(value)


def _apply_worksheet_formatting(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    widths: dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            if value is None:
                continue
            widths[idx] = max(widths.get(idx, 0), len(str(value)))

    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 10), 45)


def _get_envoi_id_from_request(request) -> int | None:
    raw = None
    try:
        raw = request.query_params.get("envoi_id") or request.headers.get("X-Envoi-Id")
    except Exception:  # noqa: BLE001
        raw = None
    if raw is None or str(raw).strip() == "":
        return None
    try:
        return int(str(raw))
    except ValueError as exc:
        raise ValidationError({"envoi_id": "Identifiant d'envoi invalide."}) from exc


def get_envoi_from_request(request, *, required: bool = True) -> Envoi | None:
    cached = getattr(request, "_cached_envoi", None)
    if cached is not None:
        return cached

    envoi_id = _get_envoi_id_from_request(request)
    if envoi_id is None:
        if required:
            raise ValidationError({"envoi_id": "Envoi requis (envoi_id)."})
        return None

    try:
        envoi = Envoi.objects.get(pk=envoi_id)
    except Envoi.DoesNotExist as exc:
        raise ValidationError({"envoi_id": "Envoi introuvable."}) from exc

    setattr(request, "_cached_envoi", envoi)
    return envoi

_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
_R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _zip_join(base: str, target: str) -> str | None:
    target_raw = str(target or "")
    if not target_raw:
        return None
    if "://" in target_raw:
        return None

    is_absolute = target_raw.startswith("/")
    target_norm = target_raw.lstrip("/")
    if not target_norm:
        return None

    if is_absolute:
        joined = posixpath.normpath(target_norm)
    else:
        joined = posixpath.normpath(posixpath.join(base, target_norm))
    if joined.startswith("../") or joined == "..":
        return None
    joined = joined.lstrip("/")
    if joined.startswith("../") or joined == "..":
        return None
    return joined


def _extract_xlsx_images_by_row(xlsx_bytes: bytes, worksheet: object) -> dict[int, list[tuple[int, bytes]]]:
    images_by_row: dict[int, list[tuple[int, bytes]]] = defaultdict(list)
    if not xlsx_bytes:
        return images_by_row

    try:
        with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as zf:
            names = set(zf.namelist())

            drawing_targets: list[str] = []
            for rel in getattr(worksheet, "_rels", []) or []:
                rel_type = str(getattr(rel, "Type", "") or "")
                target = getattr(rel, "Target", None) or getattr(rel, "target", None)
                if not target:
                    continue
                target_str = str(target)
                if rel_type.endswith("/drawing") or "drawings/" in target_str:
                    drawing_targets.append(target_str)

            ns = {"xdr": _XDR_NS, "a": _A_NS, "r": _R_NS}

            for drawing_target in drawing_targets:
                drawing_path = str(drawing_target).lstrip("/")
                if drawing_path.startswith("../"):
                    drawing_path = posixpath.normpath(posixpath.join("xl/worksheets", drawing_path))
                drawing_path = drawing_path.lstrip("/")
                if drawing_path not in names and f"xl/{drawing_path}" in names:
                    drawing_path = f"xl/{drawing_path}"
                if drawing_path not in names:
                    continue

                try:
                    drawing_xml = zf.read(drawing_path)
                except KeyError:
                    continue

                rels_path = posixpath.join(
                    posixpath.dirname(drawing_path),
                    "_rels",
                    posixpath.basename(drawing_path) + ".rels",
                )
                relmap: dict[str, str] = {}
                if rels_path in names:
                    try:
                        rels_root = ET.fromstring(zf.read(rels_path))
                    except ET.ParseError:
                        rels_root = None
                    if rels_root is not None:
                        for rel_el in rels_root.findall(f"{{{_REL_NS}}}Relationship"):
                            rid = rel_el.attrib.get("Id")
                            rel_target = rel_el.attrib.get("Target")
                            target_mode = rel_el.attrib.get("TargetMode")
                            if not rid or not rel_target or target_mode == "External":
                                continue
                            joined = _zip_join(posixpath.dirname(drawing_path), rel_target)
                            if joined:
                                relmap[rid] = joined

                try:
                    draw_root = ET.fromstring(drawing_xml)
                except ET.ParseError:
                    continue

                anchors: list[ET.Element] = []
                anchors.extend(draw_root.findall("xdr:oneCellAnchor", ns))
                anchors.extend(draw_root.findall("xdr:twoCellAnchor", ns))
                for anchor in anchors:
                    from_el = anchor.find("xdr:from", ns)
                    if from_el is None:
                        continue
                    col_el = from_el.find("xdr:col", ns)
                    row_el = from_el.find("xdr:row", ns)
                    if col_el is None or row_el is None:
                        continue
                    try:
                        col_num = int(col_el.text or "0") + 1
                        row_num = int(row_el.text or "0") + 1
                    except ValueError:
                        continue

                    blip = anchor.find(".//a:blip", ns)
                    if blip is None:
                        continue
                    rid = blip.attrib.get(f"{{{_R_NS}}}embed") or blip.attrib.get(f"{{{_R_NS}}}link")
                    if not rid:
                        continue

                    image_path = relmap.get(rid)
                    if not image_path or image_path not in names:
                        continue

                    try:
                        img_bytes = zf.read(image_path)
                    except KeyError:
                        continue
                    if img_bytes:
                        images_by_row[row_num].append((col_num, img_bytes))
    except zipfile.BadZipFile:
        return images_by_row

    return images_by_row


class HealthView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):  # noqa: ARG002
        return Response({"status": "ok"})


class MeView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        display_name = f"{user.first_name} {user.last_name}".strip() or user.get_username()
        return Response(
            {
                "id": user.id,
                "username": user.get_username(),
                "first_name": user.first_name,
                "last_name": user.last_name,
                "display_name": display_name,
                "is_staff": bool(getattr(user, "is_staff", False)),
                "is_superuser": bool(getattr(user, "is_superuser", False)),
            }
        )


class EnvoiViewSet(viewsets.ModelViewSet):
    queryset = Envoi.objects.all()
    serializer_class = EnvoiSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        if getattr(self, "action", None) in ("list", "retrieve"):
            return [permissions.IsAuthenticated()]
        return [permissions.IsAdminUser()]

    def perform_create(self, serializer):
        instance = serializer.save()
        log_audit_event(
            self.request,
            action=AuditEvent.Action.CREATE,
            entity="envoi",
            obj=instance,
            message=f"Création envoi {instance.nom}",
            envoi=instance,
        )

    def perform_update(self, serializer):
        instance = serializer.save()
        log_audit_event(
            self.request,
            action=AuditEvent.Action.UPDATE,
            entity="envoi",
            obj=instance,
            message=f"Modification envoi {instance.nom}",
            envoi=instance,
        )

    def perform_destroy(self, instance: Envoi):
        produit_ids = list(Produit.objects.filter(envoi_id=instance.id).values_list("id", flat=True))
        product_count = len(produit_ids)
        tx_count = Transaction.objects.filter(produit_id__in=produit_ids).count() if produit_ids else 0
        debt_count = Dette.objects.filter(produit_id__in=produit_ids).count() if produit_ids else 0

        with db_transaction.atomic():
            with disable_stock_recalc():
                if produit_ids:
                    Dette.objects.filter(produit_id__in=produit_ids).delete()
                    Transaction.objects.filter(produit_id__in=produit_ids).delete()
                    Produit.objects.filter(id__in=produit_ids).delete()
                instance.delete()

        log_audit_event(
            self.request,
            action=AuditEvent.Action.DELETE,
            entity="envoi",
            obj=instance,
            message=f"Suppression envoi {instance.nom}",
            metadata={
                "deleted_products": product_count,
                "deleted_transactions": tx_count,
                "deleted_debts": debt_count,
            },
            envoi=instance,
        )


class ProduitViewSet(viewsets.ModelViewSet):
    queryset = Produit.objects.all().select_related("stock", "envoi")
    serializer_class = ProduitSerializer

    def get_queryset(self):
        envoi = get_envoi_from_request(self.request, required=True)
        return super().get_queryset().filter(envoi_id=envoi.id)

    def perform_create(self, serializer):
        envoi = get_envoi_from_request(self.request, required=True)
        instance = serializer.save(envoi=envoi)
        log_audit_event(
            self.request,
            action=AuditEvent.Action.CREATE,
            entity="produit",
            obj=instance,
            envoi=envoi,
            message=f"Création produit {instance.nom}",
        )

    def perform_update(self, serializer):
        instance = serializer.save()
        envoi = getattr(instance, "envoi", None)
        log_audit_event(
            self.request,
            action=AuditEvent.Action.UPDATE,
            entity="produit",
            obj=instance,
            envoi=envoi,
            message=f"Modification produit {instance.nom}",
        )

    @action(
        detail=False,
        methods=["delete"],
        permission_classes=[permissions.IsAdminUser],
        url_path="purge",
    )
    def purge(self, request):
        envoi = get_envoi_from_request(request, required=True)
        produit_ids = list(Produit.objects.filter(envoi_id=envoi.id).values_list("id", flat=True))
        product_count = len(produit_ids)
        transaction_count = (
            Transaction.objects.filter(produit_id__in=produit_ids).count() if produit_ids else 0
        )
        debt_count = Dette.objects.filter(produit_id__in=produit_ids).count() if produit_ids else 0

        with db_transaction.atomic():
            with disable_stock_recalc():
                if produit_ids:
                    Dette.objects.filter(produit_id__in=produit_ids).delete()
                    Transaction.objects.filter(produit_id__in=produit_ids).delete()
                    Produit.objects.filter(id__in=produit_ids).delete()

        log_audit_event(
            request,
            action=AuditEvent.Action.PURGE,
            entity="produit",
            message=f"Suppression de tous les produits ({envoi.nom})",
            metadata={
                "deleted_products": product_count,
                "deleted_transactions": transaction_count,
                "deleted_debts": debt_count,
            },
            envoi=envoi,
        )

        return Response(
            {
                "deleted_products": product_count,
                "deleted_transactions": transaction_count,
                "deleted_debts": debt_count,
            }
        )

    def perform_destroy(self, instance: Produit):
        tx_count = Transaction.objects.filter(produit_id=instance.id).count()
        debt_count = Dette.objects.filter(produit_id=instance.id).count()
        envoi = getattr(instance, "envoi", None)
        log_audit_event(
            self.request,
            action=AuditEvent.Action.DELETE,
            entity="produit",
            obj=instance,
            envoi=envoi,
            message=f"Suppression produit {instance.nom}",
            metadata={
                "deleted_transactions": tx_count,
                "deleted_debts": debt_count,
            },
        )
        with db_transaction.atomic():
            Dette.objects.filter(produit_id=instance.id).delete()
            Transaction.objects.filter(produit_id=instance.id).delete()
            instance.delete()


class StockViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Stock.objects.all().select_related("produit", "produit__envoi")
    serializer_class = StockSerializer

    def get_queryset(self):
        envoi = get_envoi_from_request(self.request, required=True)
        return super().get_queryset().filter(produit__envoi_id=envoi.id)


class TransactionViewSet(viewsets.ModelViewSet):
    queryset = Transaction.objects.filter(
        type_transaction__in=(
            Transaction.TypeTransaction.ACHAT,
            Transaction.TypeTransaction.VENTE,
        )
    ).select_related("produit", "produit__envoi")
    serializer_class = TransactionSerializer

    def get_queryset(self):
        envoi = get_envoi_from_request(self.request, required=True)
        return super().get_queryset().filter(produit__envoi_id=envoi.id)

    def perform_create(self, serializer):
        envoi = get_envoi_from_request(self.request, required=True)
        produit = serializer.validated_data.get("produit")
        if produit is None or getattr(produit, "envoi_id", None) != envoi.id:
            raise ValidationError({"produit": "Ce produit n'appartient pas à l'envoi sélectionné."})
        instance = serializer.save()
        log_audit_event(
            self.request,
            action=AuditEvent.Action.CREATE,
            entity="transaction",
            obj=instance,
            message=f"{instance.type_transaction} {instance.produit.nom} x{instance.quantite}",
            envoi=envoi,
        )

    def perform_update(self, serializer):
        instance = serializer.save()
        envoi = getattr(getattr(instance, "produit", None), "envoi", None)
        log_audit_event(
            self.request,
            action=AuditEvent.Action.UPDATE,
            entity="transaction",
            obj=instance,
            message=f"{instance.type_transaction} {instance.produit.nom} x{instance.quantite}",
            envoi=envoi,
        )

    def perform_destroy(self, instance):
        envoi = getattr(getattr(instance, "produit", None), "envoi", None)
        log_audit_event(
            self.request,
            action=AuditEvent.Action.DELETE,
            entity="transaction",
            obj=instance,
            message=f"Suppression {instance.type_transaction} {instance.produit.nom} x{instance.quantite}",
            envoi=envoi,
        )
        super().perform_destroy(instance)


class TauxChangeViewSet(viewsets.ModelViewSet):
    queryset = TauxChange.objects.all()
    serializer_class = TauxChangeSerializer

    @action(detail=False, methods=["get"])
    def current(self, request):  # noqa: ARG002
        taux = get_current_exchange_rate()
        return Response({"taux_euro_cfa": str(taux) if taux is not None else None})

    def perform_create(self, serializer):
        instance = serializer.save(utilisateur=self.request.user)
        log_audit_event(
            self.request,
            action=AuditEvent.Action.CREATE,
            entity="taux_change",
            obj=instance,
            message=f"Taux EUR->CFA = {instance.taux_euro_cfa} ({instance.date_application})",
        )

    def perform_update(self, serializer):
        instance = serializer.save()
        log_audit_event(
            self.request,
            action=AuditEvent.Action.UPDATE,
            entity="taux_change",
            obj=instance,
            message=f"Taux EUR->CFA = {instance.taux_euro_cfa} ({instance.date_application})",
        )

    def perform_destroy(self, instance):
        log_audit_event(
            self.request,
            action=AuditEvent.Action.DELETE,
            entity="taux_change",
            obj=instance,
            message=f"Suppression taux EUR->CFA ({instance.taux_euro_cfa} / {instance.date_application})",
        )
        super().perform_destroy(instance)


class DetteViewSet(viewsets.ModelViewSet):
    queryset = Dette.objects.all().select_related(
        "produit", "produit__envoi", "transaction_pret", "transaction_retour"
    )
    serializer_class = DetteSerializer

    def get_queryset(self):
        envoi = get_envoi_from_request(self.request, required=True)
        return super().get_queryset().filter(produit__envoi_id=envoi.id)

    def perform_create(self, serializer):
        envoi = get_envoi_from_request(self.request, required=True)
        produit = serializer.validated_data.get("produit")
        if produit is None or getattr(produit, "envoi_id", None) != envoi.id:
            raise ValidationError({"produit": "Ce produit n'appartient pas à l'envoi sélectionné."})
        instance = serializer.save()
        log_audit_event(
            self.request,
            action=AuditEvent.Action.CREATE,
            entity="dette",
            obj=instance,
            message=f"Dette client {instance.client} x{instance.quantite_pretee} ({instance.produit.nom})",
            envoi=envoi,
        )

    def perform_update(self, serializer):
        instance = serializer.save()
        envoi = getattr(getattr(instance, "produit", None), "envoi", None)
        log_audit_event(
            self.request,
            action=AuditEvent.Action.UPDATE,
            entity="dette",
            obj=instance,
            message=f"Dette client {instance.client} x{instance.quantite_pretee} ({instance.produit.nom})",
            metadata={
                "statut": instance.statut,
                "date_retour_effective": (
                    instance.date_retour_effective.isoformat()
                    if instance.date_retour_effective is not None
                    else None
                ),
            },
            envoi=envoi,
        )

    def perform_destroy(self, instance: Dette):
        envoi = getattr(getattr(instance, "produit", None), "envoi", None)
        log_audit_event(
            self.request,
            action=AuditEvent.Action.DELETE,
            entity="dette",
            obj=instance,
            message=f"Suppression dette {instance.client} x{instance.quantite_pretee} ({instance.produit.nom})",
            envoi=envoi,
        )
        tx_retour = instance.transaction_retour
        tx_pret = instance.transaction_pret
        instance.transaction_retour = None
        instance.transaction_pret = None
        instance.save(update_fields=["transaction_retour", "transaction_pret"])
        instance.delete()

        if tx_retour:
            tx_retour.delete()
        if tx_pret:
            tx_pret.delete()


class AuditEventViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditEvent.objects.all().select_related("user", "envoi")
    serializer_class = AuditEventSerializer
    permission_classes = [permissions.IsAdminUser]

    def get_queryset(self):
        qs = super().get_queryset()

        envoi_id = _get_envoi_id_from_request(self.request)
        if envoi_id is not None:
            qs = qs.filter(envoi_id=envoi_id)

        after_id = self.request.query_params.get("after_id")
        if after_id:
            try:
                after_int = int(after_id)
            except ValueError:
                return qs.none()
            qs = qs.filter(id__gt=after_int).order_by("id")

        limit_raw = self.request.query_params.get("limit")
        limit = 200
        if limit_raw:
            try:
                limit = int(limit_raw)
            except ValueError:
                limit = 200
        limit = max(1, min(limit, 500))

        return qs[:limit]


class ProductImportView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        upload = request.FILES.get("file")
        if not upload:
            return Response(
                {"detail": "Fichier Excel manquant (champ 'file')."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        envoi = get_envoi_from_request(request, required=True)

        mode = (request.query_params.get("mode") or "append").strip().lower()
        if mode not in ("append", "upsert"):
            return Response(
                {"detail": "Paramètre 'mode' invalide (valeurs: append|upsert)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        def normalize_header(value: object) -> str:
            if value is None:
                return ""
            text = str(value).strip().lower()
            if not text:
                return ""
            text = text.replace("€", "euro").replace("fcfa", "cfa").replace("xof", "cfa")
            text = unicodedata.normalize("NFKD", text)
            text = "".join(ch for ch in text if not unicodedata.combining(ch))
            text = re.sub(r"[^a-z0-9]+", "_", text)
            text = text.strip("_")
            parts = ["euro" if part == "eur" else part for part in text.split("_") if part]
            return "_".join(parts)

        def parse_decimal(value: object) -> Decimal | None:
            if value is None:
                return None
            if isinstance(value, Decimal):
                return value
            if isinstance(value, (int, float)):
                return Decimal(str(value))

            raw = str(value).strip()
            if raw == "":
                return None

            raw = raw.replace(" ", "").replace("\u00a0", "").replace(",", ".")
            raw = re.sub(r"[^0-9.\-]+", "", raw)
            if raw in ("", ".", "-", "-."):
                return None
            return Decimal(raw)

        def parse_int(value: object) -> int | None:
            if value is None:
                return None
            if isinstance(value, bool):
                return None
            if isinstance(value, int):
                return value
            if isinstance(value, Decimal):
                if value == value.to_integral_value():
                    return int(value)
                raise ValueError
            if isinstance(value, float):
                if value.is_integer():
                    return int(value)
                raise ValueError

            raw = str(value).strip()
            if raw == "":
                return None
            raw = raw.replace(" ", "").replace("\u00a0", "")
            raw = re.sub(r"[^0-9\\-]+", "", raw)
            if raw in ("", "-"):
                return None
            return int(raw)

        def sniff_image_extension(data: bytes) -> str | None:
            if not data:
                return None
            if data.startswith(b"\x89PNG\r\n\x1a\n"):
                return "png"
            if data.startswith(b"\xff\xd8"):
                return "jpg"
            if data[:6] in (b"GIF87a", b"GIF89a"):
                return "gif"
            if data.startswith(b"RIFF") and data[8:12] == b"WEBP":
                return "webp"
            return None

        try:
            xlsx_bytes = upload.read()
        except Exception:
            return Response(
                {"detail": "Impossible de lire le fichier Excel."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not xlsx_bytes:
            return Response(
                {"detail": "Fichier Excel vide."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        except Exception:
            return Response(
                {"detail": "Impossible de lire le fichier Excel (format invalide)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                return Response(
                    {"detail": "Fichier Excel vide."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            headers = [normalize_header(cell) for cell in header_row]
            header_index = {h: i for i, h in enumerate(headers) if h}
            original_headers = ["" if cell is None else str(cell) for cell in header_row]

            images_by_row = _extract_xlsx_images_by_row(xlsx_bytes, ws)
            images_found_drawing_xml = sum(len(v) for v in images_by_row.values())
            images_found_openpyxl = 0
            try:
                from openpyxl.utils.cell import (
                    column_index_from_string,
                    coordinate_from_string,
                )

                for img in getattr(ws, "_images", []) or []:
                    try:
                        anchor = getattr(img, "anchor", None)
                        row_num: int | None = None
                        col_num: int | None = None

                        if isinstance(anchor, str):
                            col_letter, row_num = coordinate_from_string(anchor)
                            col_num = column_index_from_string(col_letter)
                        else:
                            marker = getattr(anchor, "_from", None) or getattr(anchor, "from_", None)
                            if marker is not None:
                                row = getattr(marker, "row", None)
                                col = getattr(marker, "col", None)
                                if row is not None and col is not None:
                                    row_num = int(row) + 1
                                    col_num = int(col) + 1

                        if row_num is None or col_num is None:
                            continue
                        data = img._data()
                        if not data:
                            continue
                        images_by_row[row_num].append((col_num, data))
                        images_found_openpyxl += 1
                    except Exception:  # noqa: BLE001
                        continue
            except Exception:  # noqa: BLE001
                images_found_openpyxl = 0

            if images_by_row:
                for row_num, imgs in list(images_by_row.items()):
                    seen: set[tuple[int, int, bytes, bytes]] = set()
                    deduped: list[tuple[int, bytes]] = []
                    for col_num, data in imgs:
                        if not data:
                            continue
                        key = (col_num, len(data), data[:16], data[-16:])
                        if key in seen:
                            continue
                        seen.add(key)
                        deduped.append((col_num, data))
                    images_by_row[row_num] = deduped

            def cell_value(row, names: tuple[str, ...]) -> object | None:
                for name in names:
                    idx = header_index.get(name)
                    if idx is None or idx >= len(row):
                        continue
                    value = row[idx]
                    if value is None:
                        continue
                    if isinstance(value, str) and value.strip() == "":
                        continue
                    return value
                return None

            name_headers = (
                "nom",
                "produit",
                "nom_produit",
                "nom_du_produit",
                "product",
                "designation",
                "article",
                "libelle",
            )
            carac_headers = (
                "caracteristiques",
                "caracteristique",
                "description",
                "details",
                "specifications",
            )
            category_headers = ("categorie", "category", "cat", "famille")
            image_url_headers = ("image_url", "url", "lien_image", "url_image")
            image_file_headers = ("image", "photo", "image_file", "fichier_image", "image_fichier")

            pau_euro_headers = (
                "prix_achat_unitaire_euro",
                "prix_achat_euro",
                "prix_achat",
                "pau_euro",
                "pau",
            )
            pau_cfa_headers = ("prix_achat_unitaire_cfa", "prix_achat_cfa", "pau_cfa")

            pvu_cfa_headers = (
                "prix_vente_unitaire_cfa",
                "prix_vente_cfa",
                "prix_vente",
                "pvu_cfa",
                "pvu",
            )
            pvu_euro_headers = ("prix_vente_unitaire_euro", "prix_vente_euro", "pvu_euro", "pvu_eur")

            quantite_headers = (
                "quantite",
                "quantite_achetee",
                "quantite_achete",
                "quantite_initiale",
                "quantite_initial",
                "qte",
                "qte_achetee",
                "qte_achete",
                "stock",
                "stock_initial",
                "qualite",
            )

            created = 0
            updated = 0
            skipped = 0
            merged = 0
            images_imported = 0
            errors: list[dict[str, object]] = []

            taux = get_current_exchange_rate()
            zip_images: dict[str, bytes] = {}
            zip_total_files = 0
            zip_upload = request.FILES.get("images_zip")
            if zip_upload:
                try:
                    with zipfile.ZipFile(zip_upload) as zf:
                        for info in zf.infolist():
                            if info.is_dir():
                                continue
                            zip_total_files += 1
                            if info.file_size <= 0:
                                continue
                            # Prend uniquement le nom de fichier (anti zip-slip)
                            name = info.filename.replace("\\", "/").split("/")[-1]
                            if not name:
                                continue
                            with zf.open(info) as fp:
                                data = fp.read()
                            if not data:
                                continue
                            zip_images[name.lower()] = data
                except zipfile.BadZipFile:
                    return Response(
                        {"detail": "Zip d'images invalide (champ 'images_zip')."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            def detect_column(
                keys: tuple[str, ...],
                *,
                include_any: tuple[str, ...] = (),
                exclude_any: tuple[str, ...] = (),
            ) -> tuple[str | None, int | None]:
                for key in keys:
                    idx = header_index.get(key)
                    if idx is not None:
                        return key, idx

                if not include_any:
                    return None, None

                for key, idx in header_index.items():
                    parts = key.split("_")

                    if exclude_any and any(
                        part.startswith(excluded) for excluded in exclude_any for part in parts
                    ):
                        continue
                    if include_any and not any(
                        part.startswith(included) for included in include_any for part in parts
                    ):
                        continue
                    return key, idx

                return None, None

            image_url_header, image_url_idx = detect_column(
                image_url_headers,
                include_any=("url", "lien"),
            )
            image_file_header, image_file_idx = detect_column(
                image_file_headers,
                include_any=("image", "photo"),
                exclude_any=("url",),
            )
            image_col = None
            if image_file_idx is not None:
                image_col = image_file_idx + 1
            elif image_url_idx is not None:
                image_col = image_url_idx + 1

            quantite_header, quantite_idx = detect_column(
                quantite_headers,
                include_any=("quantite", "qte", "qualite", "stock"),
                exclude_any=("vend", "vente", "restant", "dette", "pret", "pretee"),
            )

            images_found = sum(len(v) for v in images_by_row.values())
            images_rows_preview = sorted(images_by_row.keys())[:12]

            for row_number, row in enumerate(rows_iter, start=2):
                if not row or all(
                    (cell is None) or (isinstance(cell, str) and cell.strip() == "")
                    for cell in row
                ):
                    skipped += 1
                    continue

                nom = cell_value(row, name_headers)
                if nom is None or str(nom).strip() == "":
                    errors.append({"row": row_number, "field": "nom", "message": "Nom manquant."})
                    continue
                nom_str = str(nom).strip()

                data: dict[str, object] = {"nom": nom_str}

                try:
                    quantite_value: object | None
                    if quantite_idx is not None and quantite_idx < len(row):
                        quantite_value = row[quantite_idx]
                    else:
                        quantite_value = cell_value(row, quantite_headers)
                    achat_quantite = parse_int(quantite_value)
                except ValueError:
                    errors.append(
                        {
                            "row": row_number,
                            "field": "quantite",
                            "message": "Quantité invalide.",
                        }
                    )
                    continue
                if achat_quantite is not None and achat_quantite < 0:
                    errors.append(
                        {
                            "row": row_number,
                            "field": "quantite",
                            "message": "Quantité invalide (doit être >= 0).",
                        }
                    )
                    continue

                categorie = cell_value(row, category_headers)
                if categorie is not None and str(categorie).strip() != "":
                    data["categorie"] = str(categorie).strip()

                carac = cell_value(row, carac_headers)
                if carac is not None and str(carac).strip() != "":
                    data["caracteristiques"] = str(carac).strip()

                image_url = cell_value(row, image_url_headers)
                if image_url is not None:
                    image_url_str = str(image_url).strip()
                    if image_url_str:
                        # Colonne "Image URL": on accepte uniquement une vraie URL (http(s)://...).
                        # Pour importer un fichier image, utilisez la colonne "Image" (image insérée ou zip).
                        if re.match(r"^[a-z][a-z0-9+.-]*://", image_url_str, flags=re.IGNORECASE):
                            data["image_url"] = image_url_str

                image_filename_hint: str | None = None
                image_filename_cell = cell_value(row, image_file_headers)
                if image_filename_cell is not None:
                    image_filename_hint = str(image_filename_cell).strip()
                    if not image_filename_hint:
                        image_filename_hint = None

                try:
                    pau_euro = parse_decimal(cell_value(row, pau_euro_headers))
                    pau_cfa = parse_decimal(cell_value(row, pau_cfa_headers))
                    if pau_euro is None and pau_cfa is not None and taux not in (None, 0):
                        pau_euro = (pau_cfa / taux).quantize(Decimal("0.01"))
                    if pau_euro is not None:
                        data["prix_achat_unitaire_euro"] = pau_euro.quantize(Decimal("0.01"))
                except InvalidOperation:
                    errors.append(
                        {"row": row_number, "field": "prix_achat_unitaire_euro", "message": "PAU invalide."}
                    )
                    continue

                try:
                    pvu_cfa = parse_decimal(cell_value(row, pvu_cfa_headers))
                    if pvu_cfa is None:
                        pvu_euro = parse_decimal(cell_value(row, pvu_euro_headers))
                        if pvu_euro is not None:
                            if taux is None or taux == 0:
                                errors.append(
                                    {
                                        "row": row_number,
                                        "field": "prix_vente_unitaire_cfa",
                                        "message": "PVU (€) fourni mais taux EUR→CFA introuvable.",
                                    }
                                )
                                continue
                            pvu_cfa = (pvu_euro * taux).quantize(Decimal("0.01"))
                    if pvu_cfa is not None:
                        data["prix_vente_unitaire_cfa"] = pvu_cfa.quantize(Decimal("0.01"))
                except InvalidOperation:
                    errors.append(
                        {"row": row_number, "field": "prix_vente_unitaire_cfa", "message": "PVU invalide."}
                    )
                    continue

                try:
                    with db_transaction.atomic():
                        produit = None
                        merged_this_row = 0

                        if mode == "append":
                            serializer = ProduitSerializer(data=data)
                            if not serializer.is_valid():
                                errors.append({"row": row_number, "errors": serializer.errors})
                                continue
                            produit = serializer.save(envoi=envoi)
                            created += 1
                        else:
                            existing_qs = Produit.objects.filter(nom=nom_str, envoi_id=envoi.id).order_by("id")
                            existing_count = existing_qs.count()
                            if existing_count == 0:
                                serializer = ProduitSerializer(data=data)
                                if not serializer.is_valid():
                                    errors.append({"row": row_number, "errors": serializer.errors})
                                    continue
                                produit = serializer.save(envoi=envoi)
                                created += 1
                            else:
                                primary = existing_qs.first()
                                if primary is None:
                                    serializer = ProduitSerializer(data=data)
                                    if not serializer.is_valid():
                                        errors.append({"row": row_number, "errors": serializer.errors})
                                        continue
                                    produit = serializer.save(envoi=envoi)
                                    created += 1
                                else:
                                    duplicate_ids = list(existing_qs.values_list("id", flat=True)[1:])
                                    if duplicate_ids:
                                        Transaction.objects.filter(produit_id__in=duplicate_ids).update(
                                            produit_id=primary.id
                                        )
                                        Dette.objects.filter(produit_id__in=duplicate_ids).update(
                                            produit_id=primary.id
                                        )
                                        Produit.objects.filter(id__in=duplicate_ids).delete()
                                        merged_this_row = len(duplicate_ids)
                                        merged += merged_this_row

                                    serializer = ProduitSerializer(primary, data=data, partial=True)
                                    if not serializer.is_valid():
                                        errors.append({"row": row_number, "errors": serializer.errors})
                                        continue
                                    produit = serializer.save()
                                    updated += 1

                        embedded_image: bytes | None = None
                        embedded_image_name: str | None = None
                        if produit is not None and images_by_row:
                            candidates = images_by_row.get(row_number) or []
                            if candidates:
                                if image_col is not None:
                                    candidates = sorted(
                                        candidates,
                                        key=lambda item: abs(item[0] - image_col),
                                    )
                                embedded_image = candidates[0][1]
                        if (
                            embedded_image is None
                            and produit is not None
                            and zip_images
                            and image_filename_hint is not None
                        ):
                            # Nom de fichier présent dans la colonne "Image" + zip fourni
                            safe_name = image_filename_hint.replace("\\", "/").split("/")[-1]
                            if safe_name:
                                embedded_image = zip_images.get(safe_name.lower())
                                embedded_image_name = safe_name

                        if (
                            produit is not None
                            and embedded_image is not None
                            and getattr(produit, "image", None) is not None
                        ):
                            ext = sniff_image_extension(embedded_image)
                            if ext:
                                if embedded_image_name:
                                    safe_stem = re.sub(
                                        r"[^a-zA-Z0-9]+",
                                        "_",
                                        embedded_image_name.rsplit(".", 1)[0],
                                    ).strip("_")
                                    safe_stem = safe_stem[:80] if safe_stem else "image"
                                    filename = f"import_{produit.id}_{safe_stem}.{ext}"
                                else:
                                    filename = f"import_{produit.id}_{row_number}.{ext}"
                                produit.image.save(
                                    filename,
                                    ContentFile(embedded_image),
                                    save=True,
                                )
                                images_imported += 1

                        if achat_quantite is not None and achat_quantite > 0 and produit is not None:
                            pau_tx = (
                                data.get("prix_achat_unitaire_euro")
                                if "prix_achat_unitaire_euro" in data
                                else getattr(produit, "prix_achat_unitaire_euro", None)
                            )

                            tx_kwargs: dict[str, object] = {
                                "produit_id": produit.id,
                                "type_transaction": Transaction.TypeTransaction.ACHAT,
                                "quantite": int(achat_quantite),
                            }
                            if pau_tx is not None:
                                pau_tx_dec = Decimal(str(pau_tx)).quantize(Decimal("0.01"))
                                tx_kwargs["prix_unitaire_euro"] = pau_tx_dec
                                if taux is not None:
                                    tx_kwargs["taux_change"] = taux
                                    tx_kwargs["prix_unitaire_cfa"] = (pau_tx_dec * taux).quantize(
                                        Decimal("0.01")
                                    )
                            Transaction.objects.create(**tx_kwargs)
                        elif merged_this_row > 0 and produit is not None:
                            # Les reassignment via QuerySet.update n'émettent pas de signaux,
                            # donc on force un recalcul du stock après fusion.
                            recalculate_stock_for_product(produit.id)
                except InvalidOperation:
                    errors.append(
                        {
                            "row": row_number,
                            "field": "quantite",
                            "message": "Quantité invalide.",
                        }
                    )
                    continue

            payload = {
                "mode": mode,
                "created": created,
                "updated": updated,
                "merged": merged,
                "skipped": skipped,
                "images_imported": images_imported,
                "detected_columns": {
                    "quantite": quantite_header,
                    "image_url": image_url_header,
                    "image": image_file_header,
                },
                "images_found": images_found,
                "images_found_drawing_xml": images_found_drawing_xml,
                "images_found_openpyxl": images_found_openpyxl,
                "images_rows_preview": images_rows_preview,
                "images_zip_files": len(zip_images),
                "images_zip_total_entries": zip_total_files,
                "headers": [
                    {
                        "index": idx + 1,
                        "original": original_headers[idx],
                        "normalized": headers[idx],
                    }
                    for idx in range(len(headers))
                    if headers[idx] or original_headers[idx]
                ],
                "errors": errors,
            }

            log_audit_event(
                request,
                action=AuditEvent.Action.IMPORT,
                entity="produit",
                message="Import Excel produits",
                metadata={
                    "mode": mode,
                    "created": created,
                    "updated": updated,
                    "merged": merged,
                    "skipped": skipped,
                    "errors": len(errors),
                    "images_imported": images_imported,
                },
                envoi=envoi,
            )

            return Response(payload)
        finally:
            try:
                wb.close()
            except Exception:  # noqa: BLE001
                pass


class StockReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        taux = get_current_exchange_rate()
        threshold_raw = request.query_params.get("low_stock_threshold", "5")
        try:
            low_stock_threshold = max(int(threshold_raw), 0)
        except ValueError:
            low_stock_threshold = 5

        envoi = get_envoi_from_request(request, required=True)

        last_sales_by_product: dict[int, dict] = {}
        last_sales_qs = (
            Transaction.objects.filter(
                type_transaction=Transaction.TypeTransaction.VENTE,
                produit__envoi_id=envoi.id,
            )
            .exclude(prix_unitaire_cfa__isnull=True, prix_unitaire_euro__isnull=True)
            .order_by("produit_id", "-date_transaction", "-id")
            .values("produit_id", "prix_unitaire_cfa", "prix_unitaire_euro", "taux_change")
        )
        for row in last_sales_qs:
            pid = row["produit_id"]
            if pid not in last_sales_by_product:
                last_sales_by_product[pid] = row

        sales_total_cfa_by_product = defaultdict(lambda: Decimal("0"))
        sales_total_euro_by_product = defaultdict(lambda: Decimal("0"))
        sales_total_euro_ok_by_product = defaultdict(lambda: True)

        for tx in Transaction.objects.filter(
            type_transaction=Transaction.TypeTransaction.VENTE,
            produit__envoi_id=envoi.id,
        ).values(
            "produit_id",
            "quantite",
            "prix_unitaire_cfa",
            "prix_unitaire_euro",
            "taux_change",
        ):
            pid = tx["produit_id"]
            qty = Decimal(tx["quantite"] or 0)
            if qty <= 0:
                continue

            prix_cfa = tx["prix_unitaire_cfa"]
            prix_euro = tx["prix_unitaire_euro"]
            rate = tx["taux_change"] or taux

            if prix_cfa is not None:
                sales_total_cfa_by_product[pid] += qty * prix_cfa
                if rate is not None and rate != 0:
                    sales_total_euro_by_product[pid] += qty * (prix_cfa / rate)
                else:
                    sales_total_euro_ok_by_product[pid] = False
                continue

            if prix_euro is not None:
                sales_total_euro_by_product[pid] += qty * prix_euro
                if rate is not None:
                    sales_total_cfa_by_product[pid] += qty * prix_euro * rate

        debts_total_cfa_by_product = defaultdict(lambda: Decimal("0"))
        debts_total_euro_by_product = defaultdict(lambda: Decimal("0"))
        debts_total_euro_ok_by_product = defaultdict(lambda: True)

        for debt in Dette.objects.filter(
            date_retour_effective__isnull=True,
            produit__envoi_id=envoi.id,
        ).values(
            "produit_id",
            "quantite_pretee",
            "transaction_pret__prix_unitaire_cfa",
            "transaction_pret__prix_unitaire_euro",
            "transaction_pret__taux_change",
        ):
            pid = debt["produit_id"]
            qty = Decimal(debt["quantite_pretee"] or 0)
            if qty <= 0:
                continue

            prix_cfa = debt["transaction_pret__prix_unitaire_cfa"]
            prix_euro = debt["transaction_pret__prix_unitaire_euro"]
            rate = debt["transaction_pret__taux_change"] or taux

            if prix_cfa is not None:
                debts_total_cfa_by_product[pid] += qty * prix_cfa
                if rate is not None and rate != 0:
                    debts_total_euro_by_product[pid] += qty * (prix_cfa / rate)
                else:
                    debts_total_euro_ok_by_product[pid] = False
                continue

            if prix_euro is not None:
                debts_total_euro_by_product[pid] += qty * prix_euro
                if rate is not None:
                    debts_total_cfa_by_product[pid] += qty * prix_euro * rate
                continue

            debts_total_euro_ok_by_product[pid] = False

        items = []
        totals_qte_achetee = 0
        totals_qte_vendue = 0
        totals_stock_restant = 0
        totals_qte_dettes = 0

        totals_valeur_achetee_euro = Decimal("0")
        totals_valeur_vendue_euro = Decimal("0")
        totals_valeur_stock_euro = Decimal("0")
        totals_valeur_dettes_euro = Decimal("0")

        totals_valeur_achetee_cfa = Decimal("0") if taux is not None else None
        totals_valeur_stock_cfa = Decimal("0") if taux is not None else None
        totals_valeur_vendue_cfa = Decimal("0")
        totals_valeur_dettes_cfa = Decimal("0")

        totals_valeur_vendue_euro_ok = True
        totals_valeur_dettes_euro_ok = True

        for stock in Stock.objects.select_related("produit", "produit__envoi").filter(
            produit__envoi_id=envoi.id
        ):
            produit = stock.produit
            image_path = ""
            try:
                if getattr(produit, "image", None) and produit.image:
                    image_path = produit.image.url
            except Exception:  # noqa: BLE001
                image_path = ""

            pau = produit.prix_achat_unitaire_euro or Decimal("0")
            qte_achetee = Decimal(stock.quantite_initial)
            qte_vendue = Decimal(stock.quantite_vendue)
            qte_restante = Decimal(stock.quantite_restante)
            qte_pretee = Decimal(stock.quantite_pretee)

            pau_cfa = None if taux is None else (pau * taux)

            pvu_cfa = produit.prix_vente_unitaire_cfa
            sale = last_sales_by_product.get(produit.id)
            if pvu_cfa is None and sale:
                if sale["prix_unitaire_cfa"] is not None:
                    pvu_cfa = sale["prix_unitaire_cfa"]
                elif sale["prix_unitaire_euro"] is not None:
                    rate = sale["taux_change"] or taux
                    if rate is not None:
                        pvu_cfa = (sale["prix_unitaire_euro"] * rate).quantize(Decimal("0.01"))

            if pvu_cfa is not None and taux is not None:
                pvu_euro = (pvu_cfa / taux).quantize(Decimal("0.01"))
            elif sale and sale["prix_unitaire_euro"] is not None:
                pvu_euro = sale["prix_unitaire_euro"]
            else:
                pvu_euro = None

            valeur_achetee_euro = pau * qte_achetee
            valeur_achetee_cfa = None if taux is None else (valeur_achetee_euro * taux)

            valeur_vendue_cfa = sales_total_cfa_by_product[produit.id]
            valeur_vendue_euro = (
                sales_total_euro_by_product[produit.id]
                if sales_total_euro_ok_by_product[produit.id]
                else None
            )

            valeur_stock_euro = pau * qte_restante
            valeur_stock_cfa = None if taux is None else (valeur_stock_euro * taux)

            valeur_dettes_cfa = debts_total_cfa_by_product[produit.id]
            valeur_dettes_euro = (
                debts_total_euro_by_product[produit.id]
                if debts_total_euro_ok_by_product[produit.id]
                else None
            )

            totals_qte_achetee += stock.quantite_initial
            totals_qte_vendue += stock.quantite_vendue
            totals_stock_restant += stock.quantite_restante
            totals_qte_dettes += stock.quantite_pretee

            totals_valeur_achetee_euro += valeur_achetee_euro
            totals_valeur_stock_euro += valeur_stock_euro
            if valeur_vendue_euro is not None:
                totals_valeur_vendue_euro += valeur_vendue_euro
            elif stock.quantite_vendue:
                totals_valeur_vendue_euro_ok = False
            if valeur_dettes_euro is not None:
                totals_valeur_dettes_euro += valeur_dettes_euro
            elif stock.quantite_pretee:
                totals_valeur_dettes_euro_ok = False

            if totals_valeur_achetee_cfa is not None and valeur_achetee_cfa is not None:
                totals_valeur_achetee_cfa += valeur_achetee_cfa
            if totals_valeur_stock_cfa is not None and valeur_stock_cfa is not None:
                totals_valeur_stock_cfa += valeur_stock_cfa
            totals_valeur_vendue_cfa += valeur_vendue_cfa
            totals_valeur_dettes_cfa += valeur_dettes_cfa

            items.append(
                {
                    "produit_id": produit.id,
                    "nom": produit.nom,
                    "image": image_path,
                    "image_url": produit.image_url,
                    "caracteristiques": produit.caracteristiques,
                    "pau_euro": str(pau.quantize(Decimal("0.01"))),
                    "pau_cfa": None
                    if pau_cfa is None
                    else str(pau_cfa.quantize(Decimal("0.01"))),
                    "pvu_cfa": None
                    if pvu_cfa is None
                    else str(Decimal(pvu_cfa).quantize(Decimal("0.01"))),
                    "pvu_euro": None
                    if pvu_euro is None
                    else str(Decimal(pvu_euro).quantize(Decimal("0.01"))),
                    "quantite_achetee": stock.quantite_initial,
                    "valeur_achetee_euro": str(valeur_achetee_euro.quantize(Decimal("0.01"))),
                    "valeur_achetee_cfa": None
                    if valeur_achetee_cfa is None
                    else str(valeur_achetee_cfa.quantize(Decimal("0.01"))),
                    "quantite_vendue": stock.quantite_vendue,
                    "valeur_vendue_euro": None
                    if valeur_vendue_euro is None
                    else str(valeur_vendue_euro.quantize(Decimal("0.01"))),
                    "valeur_vendue_cfa": str(valeur_vendue_cfa.quantize(Decimal("0.01"))),
                    "stock_restant": stock.quantite_restante,
                    "valeur_stock_euro": str(valeur_stock_euro.quantize(Decimal("0.01"))),
                    "valeur_stock_cfa": None
                    if valeur_stock_cfa is None
                    else str(valeur_stock_cfa.quantize(Decimal("0.01"))),
                    "quantite_pretee": stock.quantite_pretee,
                    "valeur_dettes_euro": None
                    if valeur_dettes_euro is None
                    else str(valeur_dettes_euro.quantize(Decimal("0.01"))),
                    "valeur_dettes_cfa": str(valeur_dettes_cfa.quantize(Decimal("0.01"))),
                    "is_low_stock": stock.quantite_restante <= low_stock_threshold,
                }
            )

        return Response(
            {
                "taux_euro_cfa": None if taux is None else str(taux),
                "low_stock_threshold": low_stock_threshold,
                "items": items,
                "totals": {
                    "quantite_achetee": totals_qte_achetee,
                    "valeur_achetee_euro": str(
                        totals_valeur_achetee_euro.quantize(Decimal("0.01"))
                    ),
                    "valeur_achetee_cfa": None
                    if totals_valeur_achetee_cfa is None
                    else str(totals_valeur_achetee_cfa.quantize(Decimal("0.01"))),
                    "quantite_vendue": totals_qte_vendue,
                    "valeur_vendue_euro": None
                    if not totals_valeur_vendue_euro_ok
                    else str(totals_valeur_vendue_euro.quantize(Decimal("0.01"))),
                    "valeur_vendue_cfa": str(totals_valeur_vendue_cfa.quantize(Decimal("0.01"))),
                    "stock_restant": totals_stock_restant,
                    "valeur_stock_euro": str(
                        totals_valeur_stock_euro.quantize(Decimal("0.01"))
                    ),
                    "valeur_stock_cfa": None
                    if totals_valeur_stock_cfa is None
                    else str(totals_valeur_stock_cfa.quantize(Decimal("0.01"))),
                    "quantite_pretee": totals_qte_dettes,
                    "valeur_dettes_euro": None
                    if not totals_valeur_dettes_euro_ok
                    else str(totals_valeur_dettes_euro.quantize(Decimal("0.01"))),
                    "valeur_dettes_cfa": str(totals_valeur_dettes_cfa.quantize(Decimal("0.01"))),
                },
            }
        )


class MonthlyReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        taux = get_current_exchange_rate()
        year_raw = request.query_params.get("year")
        envoi = get_envoi_from_request(request, required=True)

        qs = Transaction.objects.filter(
            type_transaction__in=(
                Transaction.TypeTransaction.ACHAT,
                Transaction.TypeTransaction.VENTE,
            )
        ).filter(produit__envoi_id=envoi.id).order_by("date_transaction", "id")
        if year_raw:
            try:
                year = int(year_raw)
                qs = qs.filter(date_transaction__year=year)
            except ValueError:
                return Response(
                    {"detail": "Paramètre 'year' invalide (ex: 2025)."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        def month_key(dt) -> str:
            if isinstance(dt, date):
                d = dt
            else:
                d = dt.date()
            return f"{d.year:04d}-{d.month:02d}"

        def tx_total_euro(tx: dict) -> Decimal | None:
            prix_euro = tx.get("prix_unitaire_euro")
            if prix_euro is not None:
                return Decimal(tx["quantite"]) * prix_euro

            prix_cfa = tx.get("prix_unitaire_cfa")
            if prix_cfa is None:
                return None

            rate = tx.get("taux_change") or taux
            if rate is None or rate == 0:
                return None
            return Decimal(tx["quantite"]) * (prix_cfa / rate)

        def tx_total_cfa(tx: dict) -> Decimal | None:
            prix_cfa = tx.get("prix_unitaire_cfa")
            if prix_cfa is not None:
                return Decimal(tx["quantite"]) * prix_cfa

            prix_euro = tx.get("prix_unitaire_euro")
            if prix_euro is None:
                return None

            rate = tx.get("taux_change") or taux
            if rate is None:
                return None
            return Decimal(tx["quantite"]) * prix_euro * rate

        buckets = defaultdict(
            lambda: {
                "month": "",
                "achats_quantite": 0,
                "achats_total_euro": Decimal("0"),
                "achats_total_cfa": Decimal("0"),
                "ventes_quantite": 0,
                "ventes_total_euro": Decimal("0"),
                "ventes_total_cfa": Decimal("0"),
                "prets_quantite": 0,
                "retours_quantite": 0,
            }
        )

        fields = (
            "date_transaction",
            "type_transaction",
            "quantite",
            "prix_unitaire_euro",
            "prix_unitaire_cfa",
            "taux_change",
        )
        for tx in qs.values(*fields):
            m = month_key(tx["date_transaction"])
            bucket = buckets[m]
            bucket["month"] = m

            ttype = tx["type_transaction"]
            qty = int(tx["quantite"])
            total_euro = tx_total_euro(tx)
            total_cfa = tx_total_cfa(tx)

            if ttype == Transaction.TypeTransaction.ACHAT:
                bucket["achats_quantite"] += qty
                if total_euro is not None:
                    bucket["achats_total_euro"] += total_euro
                if total_cfa is not None:
                    bucket["achats_total_cfa"] += total_cfa
            elif ttype == Transaction.TypeTransaction.VENTE:
                bucket["ventes_quantite"] += qty
                if total_euro is not None:
                    bucket["ventes_total_euro"] += total_euro
                if total_cfa is not None:
                    bucket["ventes_total_cfa"] += total_cfa
        dettes_created_qs = Dette.objects.filter(produit__envoi_id=envoi.id)
        dettes_paid_qs = Dette.objects.filter(
            date_retour_effective__isnull=False,
            produit__envoi_id=envoi.id,
        )
        if year_raw:
            dettes_created_qs = dettes_created_qs.filter(date_pret__year=year)
            dettes_paid_qs = dettes_paid_qs.filter(date_retour_effective__year=year)

        for dette in dettes_created_qs.values("date_pret", "quantite_pretee"):
            m = month_key(dette["date_pret"])
            bucket = buckets[m]
            bucket["month"] = m
            bucket["prets_quantite"] += int(dette["quantite_pretee"])

        for dette in dettes_paid_qs.values("date_retour_effective", "quantite_pretee"):
            m = month_key(dette["date_retour_effective"])
            bucket = buckets[m]
            bucket["month"] = m
            bucket["retours_quantite"] += int(dette["quantite_pretee"])

        months = []
        totals = {
            "achats_quantite": 0,
            "achats_total_euro": Decimal("0"),
            "achats_total_cfa": Decimal("0"),
            "ventes_quantite": 0,
            "ventes_total_euro": Decimal("0"),
            "ventes_total_cfa": Decimal("0"),
            "prets_quantite": 0,
            "retours_quantite": 0,
        }

        for m in sorted(buckets.keys()):
            b = buckets[m]
            marge_brute_cfa = b["ventes_total_cfa"] - b["achats_total_cfa"]
            months.append(
                {
                    "month": b["month"],
                    "achats_quantite": b["achats_quantite"],
                    "achats_total_euro": str(b["achats_total_euro"].quantize(Decimal("0.01"))),
                    "achats_total_cfa": str(b["achats_total_cfa"].quantize(Decimal("0.01"))),
                    "ventes_quantite": b["ventes_quantite"],
                    "ventes_total_euro": str(b["ventes_total_euro"].quantize(Decimal("0.01"))),
                    "ventes_total_cfa": str(b["ventes_total_cfa"].quantize(Decimal("0.01"))),
                    "marge_brute_cfa": str(marge_brute_cfa.quantize(Decimal("0.01"))),
                    "prets_quantite": b["prets_quantite"],
                    "retours_quantite": b["retours_quantite"],
                }
            )

            totals["achats_quantite"] += b["achats_quantite"]
            totals["achats_total_euro"] += b["achats_total_euro"]
            totals["achats_total_cfa"] += b["achats_total_cfa"]
            totals["ventes_quantite"] += b["ventes_quantite"]
            totals["ventes_total_euro"] += b["ventes_total_euro"]
            totals["ventes_total_cfa"] += b["ventes_total_cfa"]
            totals["prets_quantite"] += b["prets_quantite"]
            totals["retours_quantite"] += b["retours_quantite"]

        totals_out = {
            "achats_quantite": totals["achats_quantite"],
            "achats_total_euro": str(totals["achats_total_euro"].quantize(Decimal("0.01"))),
            "achats_total_cfa": str(totals["achats_total_cfa"].quantize(Decimal("0.01"))),
            "ventes_quantite": totals["ventes_quantite"],
            "ventes_total_euro": str(totals["ventes_total_euro"].quantize(Decimal("0.01"))),
            "ventes_total_cfa": str(totals["ventes_total_cfa"].quantize(Decimal("0.01"))),
            "marge_brute_cfa": str(
                (totals["ventes_total_cfa"] - totals["achats_total_cfa"]).quantize(
                    Decimal("0.01")
                )
            ),
            "prets_quantite": totals["prets_quantite"],
            "retours_quantite": totals["retours_quantite"],
        }

        return Response(
            {
                "taux_euro_cfa": None if taux is None else str(taux),
                "months": months,
                "totals": totals_out,
            }
        )


class ExportTransactionsXlsxView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        envoi = get_envoi_from_request(request, required=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        ws.append(
            [
                "Date",
                "Produit",
                "Type",
                "Quantité",
                "Prix unitaire (€)",
                "Prix unitaire (CFA)",
                "Taux EUR->CFA",
                "Total (€)",
                "Total (CFA)",
                "Client/Fournisseur",
                "Notes",
            ]
        )

        for tx in (
            Transaction.objects.filter(
                type_transaction__in=(
                    Transaction.TypeTransaction.ACHAT,
                    Transaction.TypeTransaction.VENTE,
                )
            )
            .filter(produit__envoi_id=envoi.id)
            .select_related("produit")
            .order_by("-date_transaction", "-id")
        ):
            total_euro = None
            if tx.prix_unitaire_euro is not None:
                total_euro = (Decimal(tx.quantite) * tx.prix_unitaire_euro).quantize(
                    Decimal("0.01")
                )

            total_cfa = None
            if tx.prix_unitaire_cfa is not None:
                total_cfa = (Decimal(tx.quantite) * tx.prix_unitaire_cfa).quantize(Decimal("0.01"))
            elif tx.prix_unitaire_euro is not None:
                rate = tx.taux_change or get_current_exchange_rate()
                if rate is not None:
                    total_cfa = (Decimal(tx.quantite) * tx.prix_unitaire_euro * rate).quantize(
                        Decimal("0.01")
                    )

            ws.append(
                [
                    tx.date_transaction.isoformat(sep=" ", timespec="seconds"),
                    tx.produit.nom,
                    tx.type_transaction,
                    tx.quantite,
                    float(tx.prix_unitaire_euro) if tx.prix_unitaire_euro is not None else None,
                    float(tx.prix_unitaire_cfa) if tx.prix_unitaire_cfa is not None else None,
                    float(tx.taux_change) if tx.taux_change is not None else None,
                    None if total_euro is None else float(total_euro),
                    None if total_cfa is None else float(total_cfa),
                    tx.client_fournisseur,
                    tx.notes,
                ]
            )

        _apply_worksheet_formatting(ws)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="transactions.xlsx"'
        wb.save(response)
        return response


class ExportTransactionsCsvView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        envoi = get_envoi_from_request(request, required=True)
        out = io.StringIO(newline="")
        writer = csv.writer(out, delimiter=";")
        writer.writerow(
            [
                "Date",
                "Produit",
                "Type",
                "Quantité",
                "Prix unitaire (€)",
                "Prix unitaire (CFA)",
                "Taux EUR->CFA",
                "Total (€)",
                "Total (CFA)",
                "Client/Fournisseur",
                "Notes",
            ]
        )

        taux = get_current_exchange_rate()
        for tx in (
            Transaction.objects.filter(
                type_transaction__in=(
                    Transaction.TypeTransaction.ACHAT,
                    Transaction.TypeTransaction.VENTE,
                )
            )
            .filter(produit__envoi_id=envoi.id)
            .select_related("produit")
            .order_by("-date_transaction", "-id")
        ):
            total_euro = None
            if tx.prix_unitaire_euro is not None:
                total_euro = (Decimal(tx.quantite) * tx.prix_unitaire_euro).quantize(
                    Decimal("0.01")
                )

            total_cfa = None
            if tx.prix_unitaire_cfa is not None:
                total_cfa = (Decimal(tx.quantite) * tx.prix_unitaire_cfa).quantize(Decimal("0.01"))
            elif tx.prix_unitaire_euro is not None:
                rate = tx.taux_change or taux
                if rate is not None:
                    total_cfa = (Decimal(tx.quantite) * tx.prix_unitaire_euro * rate).quantize(
                        Decimal("0.01")
                    )

            writer.writerow(
                [
                    tx.date_transaction.isoformat(sep=" ", timespec="seconds"),
                    tx.produit.nom,
                    tx.type_transaction,
                    tx.quantite,
                    _csv_cell(tx.prix_unitaire_euro),
                    _csv_cell(tx.prix_unitaire_cfa),
                    _csv_cell(tx.taux_change),
                    _csv_cell(total_euro),
                    _csv_cell(total_cfa),
                    tx.client_fournisseur,
                    tx.notes,
                ]
            )

        response = HttpResponse(
            out.getvalue().encode("utf-8-sig"),
            content_type="text/csv; charset=utf-8",
        )
        response["Content-Disposition"] = 'attachment; filename="transactions.csv"'
        return response


class ExportMonthlyXlsxView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        year_raw = request.query_params.get("year")
        report = MonthlyReportView().get(request).data

        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly"

        ws.append(
            [
                "Mois",
                "Achats (qté)",
                "Achats (€)",
                "Achats (CFA)",
                "Ventes (qté)",
                "Ventes (€)",
                "Ventes (CFA)",
                "Marge brute (CFA)",
                "Dettes créées (qté)",
                "Dettes soldées (qté)",
            ]
        )

        for row in report.get("months", []):
            ws.append(
                [
                    row.get("month"),
                    row.get("achats_quantite"),
                    float(Decimal(row.get("achats_total_euro", "0")).quantize(Decimal("0.01"))),
                    float(Decimal(row.get("achats_total_cfa", "0")).quantize(Decimal("0.01"))),
                    row.get("ventes_quantite"),
                    float(Decimal(row.get("ventes_total_euro", "0")).quantize(Decimal("0.01"))),
                    float(Decimal(row.get("ventes_total_cfa", "0")).quantize(Decimal("0.01"))),
                    float(Decimal(row.get("marge_brute_cfa", "0")).quantize(Decimal("0.01"))),
                    row.get("prets_quantite"),
                    row.get("retours_quantite"),
                ]
            )

        totals = report.get("totals") or {}
        if totals:
            ws.append([])
            ws.append(
                [
                    "TOTAL",
                    totals.get("achats_quantite"),
                    float(Decimal(totals.get("achats_total_euro", "0")).quantize(Decimal("0.01"))),
                    float(Decimal(totals.get("achats_total_cfa", "0")).quantize(Decimal("0.01"))),
                    totals.get("ventes_quantite"),
                    float(Decimal(totals.get("ventes_total_euro", "0")).quantize(Decimal("0.01"))),
                    float(Decimal(totals.get("ventes_total_cfa", "0")).quantize(Decimal("0.01"))),
                    float(Decimal(totals.get("marge_brute_cfa", "0")).quantize(Decimal("0.01"))),
                    totals.get("prets_quantite"),
                    totals.get("retours_quantite"),
                ]
            )

        _apply_worksheet_formatting(ws)

        filename = "monthly.xlsx" if not year_raw else f"monthly_{year_raw}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class ExportMonthlyCsvView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        year_raw = request.query_params.get("year")
        report = MonthlyReportView().get(request).data

        def dec(value):
            if value is None:
                return ""
            return _csv_cell(Decimal(str(value)))

        out = io.StringIO(newline="")
        writer = csv.writer(out, delimiter=";")
        writer.writerow(
            [
                "Mois",
                "Achats (qté)",
                "Achats (€)",
                "Achats (CFA)",
                "Ventes (qté)",
                "Ventes (€)",
                "Ventes (CFA)",
                "Marge brute (CFA)",
                "Dettes créées (qté)",
                "Dettes soldées (qté)",
            ]
        )
        for row in report.get("months", []):
            writer.writerow(
                [
                    row.get("month"),
                    row.get("achats_quantite"),
                    dec(row.get("achats_total_euro")),
                    dec(row.get("achats_total_cfa")),
                    row.get("ventes_quantite"),
                    dec(row.get("ventes_total_euro")),
                    dec(row.get("ventes_total_cfa")),
                    dec(row.get("marge_brute_cfa")),
                    row.get("prets_quantite"),
                    row.get("retours_quantite"),
                ]
            )

        totals = report.get("totals") or {}
        if totals:
            writer.writerow([])
            writer.writerow(
                [
                    "TOTAL",
                    totals.get("achats_quantite"),
                    dec(totals.get("achats_total_euro")),
                    dec(totals.get("achats_total_cfa")),
                    totals.get("ventes_quantite"),
                    dec(totals.get("ventes_total_euro")),
                    dec(totals.get("ventes_total_cfa")),
                    dec(totals.get("marge_brute_cfa")),
                    totals.get("prets_quantite"),
                    totals.get("retours_quantite"),
                ]
            )

        filename = "monthly.csv" if not year_raw else f"monthly_{year_raw}.csv"
        response = HttpResponse(
            out.getvalue().encode("utf-8-sig"),
            content_type="text/csv; charset=utf-8",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class ExportStockXlsxView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        taux = get_current_exchange_rate()
        envoi = get_envoi_from_request(request, required=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock"

        last_sales_by_product: dict[int, dict] = {}
        last_sales_qs = (
            Transaction.objects.filter(
                type_transaction=Transaction.TypeTransaction.VENTE,
                produit__envoi_id=envoi.id,
            )
            .exclude(prix_unitaire_cfa__isnull=True, prix_unitaire_euro__isnull=True)
            .order_by("produit_id", "-date_transaction", "-id")
            .values("produit_id", "prix_unitaire_cfa", "prix_unitaire_euro", "taux_change")
        )
        for row in last_sales_qs:
            pid = row["produit_id"]
            if pid not in last_sales_by_product:
                last_sales_by_product[pid] = row

        sales_total_cfa_by_product = defaultdict(lambda: Decimal("0"))
        sales_total_euro_by_product = defaultdict(lambda: Decimal("0"))
        sales_total_euro_ok_by_product = defaultdict(lambda: True)

        for tx in Transaction.objects.filter(
            type_transaction=Transaction.TypeTransaction.VENTE,
            produit__envoi_id=envoi.id,
        ).values(
            "produit_id",
            "quantite",
            "prix_unitaire_cfa",
            "prix_unitaire_euro",
            "taux_change",
        ):
            pid = tx["produit_id"]
            qty = Decimal(tx["quantite"] or 0)
            if qty <= 0:
                continue

            prix_cfa = tx["prix_unitaire_cfa"]
            prix_euro = tx["prix_unitaire_euro"]
            rate = tx["taux_change"] or taux

            if prix_cfa is not None:
                sales_total_cfa_by_product[pid] += qty * prix_cfa
                if rate is not None and rate != 0:
                    sales_total_euro_by_product[pid] += qty * (prix_cfa / rate)
                else:
                    sales_total_euro_ok_by_product[pid] = False
                continue

            if prix_euro is not None:
                sales_total_euro_by_product[pid] += qty * prix_euro
                if rate is not None:
                    sales_total_cfa_by_product[pid] += qty * prix_euro * rate

        debts_total_cfa_by_product = defaultdict(lambda: Decimal("0"))
        debts_total_euro_by_product = defaultdict(lambda: Decimal("0"))
        debts_total_euro_ok_by_product = defaultdict(lambda: True)

        for debt in Dette.objects.filter(
            date_retour_effective__isnull=True,
            produit__envoi_id=envoi.id,
        ).values(
            "produit_id",
            "quantite_pretee",
            "transaction_pret__prix_unitaire_cfa",
            "transaction_pret__prix_unitaire_euro",
            "transaction_pret__taux_change",
        ):
            pid = debt["produit_id"]
            qty = Decimal(debt["quantite_pretee"] or 0)
            if qty <= 0:
                continue

            prix_cfa = debt["transaction_pret__prix_unitaire_cfa"]
            prix_euro = debt["transaction_pret__prix_unitaire_euro"]
            rate = debt["transaction_pret__taux_change"] or taux

            if prix_cfa is not None:
                debts_total_cfa_by_product[pid] += qty * prix_cfa
                if rate is not None and rate != 0:
                    debts_total_euro_by_product[pid] += qty * (prix_cfa / rate)
                else:
                    debts_total_euro_ok_by_product[pid] = False
                continue

            if prix_euro is not None:
                debts_total_euro_by_product[pid] += qty * prix_euro
                if rate is not None:
                    debts_total_cfa_by_product[pid] += qty * prix_euro * rate
                continue

            debts_total_euro_ok_by_product[pid] = False

        ws.append(
            [
                "Produit",
                "Caractéristiques",
                "PAU (€)",
                "PAU (CFA)",
                "PVU (CFA)",
                "PVU (€)",
                "Quantité achetée",
                "Valeur achetée (€)",
                "Valeur achetée (CFA)",
                "Quantité vendue",
                "Valeur vendue (€)",
                "Valeur vendue (CFA)",
                "Stock restant",
                "Valeur stock (€)",
                "Valeur stock (CFA)",
                "Dettes clients (qté en cours)",
                "Valeur dettes (€)",
                "Valeur dettes (CFA)",
            ]
        )

        totals_qte_achetee = 0
        totals_qte_vendue = 0
        totals_stock_restant = 0
        totals_qte_dettes = 0

        totals_valeur_achetee_euro = Decimal("0")
        totals_valeur_vendue_euro = Decimal("0")
        totals_valeur_stock_euro = Decimal("0")
        totals_valeur_dettes_euro = Decimal("0")

        totals_valeur_achetee_cfa = Decimal("0") if taux is not None else None
        totals_valeur_stock_cfa = Decimal("0") if taux is not None else None
        totals_valeur_vendue_cfa = Decimal("0")
        totals_valeur_dettes_cfa = Decimal("0")

        totals_valeur_vendue_euro_ok = True
        totals_valeur_dettes_euro_ok = True

        for stock in Stock.objects.select_related("produit", "produit__envoi").filter(
            produit__envoi_id=envoi.id
        ):
            produit = stock.produit
            pau = produit.prix_achat_unitaire_euro or Decimal("0")
            qte_achetee = Decimal(stock.quantite_initial)
            qte_vendue = Decimal(stock.quantite_vendue)
            qte_restante = Decimal(stock.quantite_restante)
            qte_pretee = Decimal(stock.quantite_pretee)

            pau_cfa = None if taux is None else (pau * taux)

            pvu_cfa = produit.prix_vente_unitaire_cfa
            sale = last_sales_by_product.get(produit.id)
            if pvu_cfa is None and sale:
                if sale["prix_unitaire_cfa"] is not None:
                    pvu_cfa = sale["prix_unitaire_cfa"]
                elif sale["prix_unitaire_euro"] is not None:
                    rate = sale["taux_change"] or taux
                    if rate is not None:
                        pvu_cfa = (sale["prix_unitaire_euro"] * rate).quantize(
                            Decimal("0.01")
                        )

            if pvu_cfa is not None and taux is not None:
                pvu_euro = (pvu_cfa / taux).quantize(Decimal("0.01"))
            elif sale and sale["prix_unitaire_euro"] is not None:
                pvu_euro = sale["prix_unitaire_euro"]
            else:
                pvu_euro = None

            valeur_achetee_euro = pau * qte_achetee
            valeur_achetee_cfa = None if taux is None else (valeur_achetee_euro * taux)

            valeur_vendue_cfa = sales_total_cfa_by_product[produit.id]
            valeur_vendue_euro = (
                sales_total_euro_by_product[produit.id]
                if sales_total_euro_ok_by_product[produit.id]
                else None
            )

            valeur_stock_euro = pau * qte_restante
            valeur_stock_cfa = None if taux is None else (valeur_stock_euro * taux)

            valeur_dettes_cfa = debts_total_cfa_by_product[produit.id]
            valeur_dettes_euro = (
                debts_total_euro_by_product[produit.id]
                if debts_total_euro_ok_by_product[produit.id]
                else None
            )

            totals_qte_achetee += stock.quantite_initial
            totals_qte_vendue += stock.quantite_vendue
            totals_stock_restant += stock.quantite_restante
            totals_qte_dettes += stock.quantite_pretee

            totals_valeur_achetee_euro += valeur_achetee_euro
            totals_valeur_stock_euro += valeur_stock_euro
            if valeur_vendue_euro is not None:
                totals_valeur_vendue_euro += valeur_vendue_euro
            elif stock.quantite_vendue:
                totals_valeur_vendue_euro_ok = False
            if valeur_dettes_euro is not None:
                totals_valeur_dettes_euro += valeur_dettes_euro
            elif stock.quantite_pretee:
                totals_valeur_dettes_euro_ok = False

            if totals_valeur_achetee_cfa is not None and valeur_achetee_cfa is not None:
                totals_valeur_achetee_cfa += valeur_achetee_cfa
            if totals_valeur_stock_cfa is not None and valeur_stock_cfa is not None:
                totals_valeur_stock_cfa += valeur_stock_cfa
            totals_valeur_vendue_cfa += valeur_vendue_cfa
            totals_valeur_dettes_cfa += valeur_dettes_cfa

            ws.append(
                [
                    produit.nom,
                    produit.caracteristiques,
                    float(pau),
                    None
                    if pau_cfa is None
                    else float(pau_cfa.quantize(Decimal("0.01"))),
                    None
                    if pvu_cfa is None
                    else float(Decimal(pvu_cfa).quantize(Decimal("0.01"))),
                    None
                    if pvu_euro is None
                    else float(Decimal(pvu_euro).quantize(Decimal("0.01"))),
                    stock.quantite_initial,
                    float(valeur_achetee_euro.quantize(Decimal("0.01"))),
                    None
                    if valeur_achetee_cfa is None
                    else float(valeur_achetee_cfa.quantize(Decimal("0.01"))),
                    stock.quantite_vendue,
                    None
                    if valeur_vendue_euro is None
                    else float(valeur_vendue_euro.quantize(Decimal("0.01"))),
                    float(valeur_vendue_cfa.quantize(Decimal("0.01"))),
                    stock.quantite_restante,
                    float(valeur_stock_euro.quantize(Decimal("0.01"))),
                    None
                    if valeur_stock_cfa is None
                    else float(valeur_stock_cfa.quantize(Decimal("0.01"))),
                    stock.quantite_pretee,
                    None
                    if valeur_dettes_euro is None
                    else float(valeur_dettes_euro.quantize(Decimal("0.01"))),
                    float(valeur_dettes_cfa.quantize(Decimal("0.01"))),
                ]
            )

        ws.append([])
        ws.append(
            [
                "TOTAL",
                "",
                None,
                None,
                None,
                None,
                totals_qte_achetee,
                float(totals_valeur_achetee_euro.quantize(Decimal("0.01"))),
                None
                if totals_valeur_achetee_cfa is None
                else float(totals_valeur_achetee_cfa.quantize(Decimal("0.01"))),
                totals_qte_vendue,
                None
                if not totals_valeur_vendue_euro_ok
                else float(totals_valeur_vendue_euro.quantize(Decimal("0.01"))),
                float(totals_valeur_vendue_cfa.quantize(Decimal("0.01"))),
                totals_stock_restant,
                float(totals_valeur_stock_euro.quantize(Decimal("0.01"))),
                None
                if totals_valeur_stock_cfa is None
                else float(totals_valeur_stock_cfa.quantize(Decimal("0.01"))),
                totals_qte_dettes,
                None
                if not totals_valeur_dettes_euro_ok
                else float(totals_valeur_dettes_euro.quantize(Decimal("0.01"))),
                float(totals_valeur_dettes_cfa.quantize(Decimal("0.01"))),
            ]
        )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="stock.xlsx"'
        _apply_worksheet_formatting(ws)
        wb.save(response)
        return response


class ExportStockCsvView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        taux = get_current_exchange_rate()
        envoi = get_envoi_from_request(request, required=True)

        last_sales_by_product: dict[int, dict] = {}
        last_sales_qs = (
            Transaction.objects.filter(
                type_transaction=Transaction.TypeTransaction.VENTE,
                produit__envoi_id=envoi.id,
            )
            .exclude(prix_unitaire_cfa__isnull=True, prix_unitaire_euro__isnull=True)
            .order_by("produit_id", "-date_transaction", "-id")
            .values("produit_id", "prix_unitaire_cfa", "prix_unitaire_euro", "taux_change")
        )
        for row in last_sales_qs:
            pid = row["produit_id"]
            if pid not in last_sales_by_product:
                last_sales_by_product[pid] = row

        sales_total_cfa_by_product = defaultdict(lambda: Decimal("0"))
        sales_total_euro_by_product = defaultdict(lambda: Decimal("0"))
        sales_total_euro_ok_by_product = defaultdict(lambda: True)

        for tx in Transaction.objects.filter(
            type_transaction=Transaction.TypeTransaction.VENTE,
            produit__envoi_id=envoi.id,
        ).values(
            "produit_id",
            "quantite",
            "prix_unitaire_cfa",
            "prix_unitaire_euro",
            "taux_change",
        ):
            pid = tx["produit_id"]
            qty = Decimal(tx["quantite"] or 0)
            if qty <= 0:
                continue

            prix_cfa = tx["prix_unitaire_cfa"]
            prix_euro = tx["prix_unitaire_euro"]
            rate = tx["taux_change"] or taux

            if prix_cfa is not None:
                sales_total_cfa_by_product[pid] += qty * prix_cfa
                if rate is not None and rate != 0:
                    sales_total_euro_by_product[pid] += qty * (prix_cfa / rate)
                else:
                    sales_total_euro_ok_by_product[pid] = False
                continue

            if prix_euro is not None:
                sales_total_euro_by_product[pid] += qty * prix_euro
                if rate is not None:
                    sales_total_cfa_by_product[pid] += qty * prix_euro * rate

        debts_total_cfa_by_product = defaultdict(lambda: Decimal("0"))
        debts_total_euro_by_product = defaultdict(lambda: Decimal("0"))
        debts_total_euro_ok_by_product = defaultdict(lambda: True)

        for debt in Dette.objects.filter(
            date_retour_effective__isnull=True,
            produit__envoi_id=envoi.id,
        ).values(
            "produit_id",
            "quantite_pretee",
            "transaction_pret__prix_unitaire_cfa",
            "transaction_pret__prix_unitaire_euro",
            "transaction_pret__taux_change",
        ):
            pid = debt["produit_id"]
            qty = Decimal(debt["quantite_pretee"] or 0)
            if qty <= 0:
                continue

            prix_cfa = debt["transaction_pret__prix_unitaire_cfa"]
            prix_euro = debt["transaction_pret__prix_unitaire_euro"]
            rate = debt["transaction_pret__taux_change"] or taux

            if prix_cfa is not None:
                debts_total_cfa_by_product[pid] += qty * prix_cfa
                if rate is not None and rate != 0:
                    debts_total_euro_by_product[pid] += qty * (prix_cfa / rate)
                else:
                    debts_total_euro_ok_by_product[pid] = False
                continue

            if prix_euro is not None:
                debts_total_euro_by_product[pid] += qty * prix_euro
                if rate is not None:
                    debts_total_cfa_by_product[pid] += qty * prix_euro * rate
                continue

            debts_total_euro_ok_by_product[pid] = False

        out = io.StringIO(newline="")
        writer = csv.writer(out, delimiter=";")
        writer.writerow(
            [
                "Produit",
                "Caractéristiques",
                "PAU (€)",
                "PAU (CFA)",
                "PVU (CFA)",
                "PVU (€)",
                "Quantité achetée",
                "Valeur achetée (€)",
                "Valeur achetée (CFA)",
                "Quantité vendue",
                "Valeur vendue (€)",
                "Valeur vendue (CFA)",
                "Stock restant",
                "Valeur stock (€)",
                "Valeur stock (CFA)",
                "Dettes clients (qté en cours)",
                "Valeur dettes (€)",
                "Valeur dettes (CFA)",
            ]
        )

        totals_qte_achetee = 0
        totals_qte_vendue = 0
        totals_stock_restant = 0
        totals_qte_dettes = 0

        totals_valeur_achetee_euro = Decimal("0")
        totals_valeur_vendue_euro = Decimal("0")
        totals_valeur_stock_euro = Decimal("0")
        totals_valeur_dettes_euro = Decimal("0")

        totals_valeur_achetee_cfa = Decimal("0") if taux is not None else None
        totals_valeur_stock_cfa = Decimal("0") if taux is not None else None
        totals_valeur_vendue_cfa = Decimal("0")
        totals_valeur_dettes_cfa = Decimal("0")

        totals_valeur_vendue_euro_ok = True
        totals_valeur_dettes_euro_ok = True

        for stock in Stock.objects.select_related("produit", "produit__envoi").filter(
            produit__envoi_id=envoi.id
        ):
            produit = stock.produit
            pau = produit.prix_achat_unitaire_euro or Decimal("0")
            qte_achetee = Decimal(stock.quantite_initial)
            qte_vendue = Decimal(stock.quantite_vendue)
            qte_restante = Decimal(stock.quantite_restante)
            qte_pretee = Decimal(stock.quantite_pretee)

            pau_cfa = None if taux is None else (pau * taux)

            pvu_cfa = produit.prix_vente_unitaire_cfa
            sale = last_sales_by_product.get(produit.id)
            if pvu_cfa is None and sale:
                if sale["prix_unitaire_cfa"] is not None:
                    pvu_cfa = sale["prix_unitaire_cfa"]
                elif sale["prix_unitaire_euro"] is not None:
                    rate = sale["taux_change"] or taux
                    if rate is not None:
                        pvu_cfa = (sale["prix_unitaire_euro"] * rate).quantize(Decimal("0.01"))

            if pvu_cfa is not None and taux is not None:
                pvu_euro = (pvu_cfa / taux).quantize(Decimal("0.01"))
            elif sale and sale["prix_unitaire_euro"] is not None:
                pvu_euro = sale["prix_unitaire_euro"]
            else:
                pvu_euro = None

            valeur_achetee_euro = pau * qte_achetee
            valeur_achetee_cfa = None if taux is None else (valeur_achetee_euro * taux)

            valeur_vendue_cfa = sales_total_cfa_by_product[produit.id]
            valeur_vendue_euro = (
                sales_total_euro_by_product[produit.id]
                if sales_total_euro_ok_by_product[produit.id]
                else None
            )

            valeur_stock_euro = pau * qte_restante
            valeur_stock_cfa = None if taux is None else (valeur_stock_euro * taux)

            valeur_dettes_cfa = debts_total_cfa_by_product[produit.id]
            valeur_dettes_euro = (
                debts_total_euro_by_product[produit.id]
                if debts_total_euro_ok_by_product[produit.id]
                else None
            )

            totals_qte_achetee += stock.quantite_initial
            totals_qte_vendue += stock.quantite_vendue
            totals_stock_restant += stock.quantite_restante
            totals_qte_dettes += stock.quantite_pretee

            totals_valeur_achetee_euro += valeur_achetee_euro
            totals_valeur_stock_euro += valeur_stock_euro
            if valeur_vendue_euro is not None:
                totals_valeur_vendue_euro += valeur_vendue_euro
            elif stock.quantite_vendue:
                totals_valeur_vendue_euro_ok = False
            if valeur_dettes_euro is not None:
                totals_valeur_dettes_euro += valeur_dettes_euro
            elif stock.quantite_pretee:
                totals_valeur_dettes_euro_ok = False

            if totals_valeur_achetee_cfa is not None and valeur_achetee_cfa is not None:
                totals_valeur_achetee_cfa += valeur_achetee_cfa
            if totals_valeur_stock_cfa is not None and valeur_stock_cfa is not None:
                totals_valeur_stock_cfa += valeur_stock_cfa

            totals_valeur_vendue_cfa += valeur_vendue_cfa
            totals_valeur_dettes_cfa += valeur_dettes_cfa

            writer.writerow(
                [
                    produit.nom,
                    produit.caracteristiques,
                    _csv_cell(pau),
                    _csv_cell(pau_cfa),
                    _csv_cell(pvu_cfa),
                    _csv_cell(pvu_euro),
                    stock.quantite_initial,
                    _csv_cell(valeur_achetee_euro),
                    _csv_cell(valeur_achetee_cfa),
                    stock.quantite_vendue,
                    _csv_cell(valeur_vendue_euro),
                    _csv_cell(valeur_vendue_cfa),
                    stock.quantite_restante,
                    _csv_cell(valeur_stock_euro),
                    _csv_cell(valeur_stock_cfa),
                    stock.quantite_pretee,
                    _csv_cell(valeur_dettes_euro),
                    _csv_cell(valeur_dettes_cfa),
                ]
            )

        writer.writerow([])
        writer.writerow(
            [
                "TOTAL",
                "",
                "",
                "",
                "",
                "",
                totals_qte_achetee,
                _csv_cell(totals_valeur_achetee_euro),
                _csv_cell(totals_valeur_achetee_cfa),
                totals_qte_vendue,
                "" if not totals_valeur_vendue_euro_ok else _csv_cell(totals_valeur_vendue_euro),
                _csv_cell(totals_valeur_vendue_cfa),
                totals_stock_restant,
                _csv_cell(totals_valeur_stock_euro),
                _csv_cell(totals_valeur_stock_cfa),
                totals_qte_dettes,
                "" if not totals_valeur_dettes_euro_ok else _csv_cell(totals_valeur_dettes_euro),
                _csv_cell(totals_valeur_dettes_cfa),
            ]
        )

        response = HttpResponse(
            out.getvalue().encode("utf-8-sig"),
            content_type="text/csv; charset=utf-8",
        )
        response["Content-Disposition"] = 'attachment; filename="stock.csv"'
        return response
